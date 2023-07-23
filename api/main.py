from datetime import datetime
from fastapi import FastAPI, UploadFile, HTTPException, Path, Query
import os
from openpyxl import load_workbook

from db.models import User, Payments, Plans, Dictionary, Credits
from main import session

app = FastAPI(
    debug=True
)


@app.get(
    "/user_credits/{user_id}",
    response_description="Повертає словник в якому ключем є id кредиту клієнта, "
                         "а значенням детальна інформація за цим кредитом",
    description="Метод повертатає список з інформацією про "
                "всі кредити клієнта за його id",
)
async def user_credits(user_id: int = Path(description="id клієнта")) -> dict:
    user = User.get(user_id)
    if not user:
        return{"response": "user not found"}
    user_credits: list[User] | None = user.credits
    response = dict()
    for credit in user_credits:
        closed = True if credit.actual_return_date else False
        credit_info = {
            "Дата видачі кредиту": credit.issuance_date,
            "Статус закриття": "Закритий" if closed else "Незакритий",
            "Сума видачі": credit.body,
            "Нараховані відсотки": credit.percent
        }
        if closed:
            credit_info["Дата повернення кредиту"] = credit.actual_return_date
        else:
            credit_info["Крайня дата повернення кредиту"] = credit.return_date
            date_diff = datetime.today().date() - credit.return_date
            credit_info[
                "Кількість днів прострочення кредиту"
            ] = date_diff.days
            credit_info[
                "Сума платежів по тілу"
            ] = Payments.get_sum_by_credit(credit.id, "Тіло")
            credit_info[
                "Сума платежів по тілу"
            ] = Payments.get_sum_by_credit(credit.id, "Відсотки")
        response[credit.id] = credit_info
    return response


@app.post("/plans_insert")
async def plans_insert(file: UploadFile):
    try:
        file_path = f"db/raw_data/{file.filename}"
        with open(file_path, "wb") as f:
            f.write(await file.read())

        wb = load_workbook(file_path)
        ws = wb.active

        if ws.cell(row=1, column=1).value != "Місяць плану" or \
                ws.cell(row=1, column=2).value != "Назва категорії плану (видача/збір)" or \
                ws.cell(row=1, column=3).value != "Сума (видача/збір)":
            os.remove(file_path)
            detail = "Неправильний формат файлу"
            raise HTTPException(status_code=400, detail=detail)

        for row in ws.iter_rows(min_row=2, values_only=True):
            month, category, amount = row
            month = month.date()
            category_id = Dictionary.get_id_by_name(category)

            if month is None or not month.strftime("%d") == "01":
                os.remove(file_path)
                detail = "Неправильний формат місяця плану, має починатися з першого числf місяця"
                raise HTTPException(status_code=400, detail=detail)

            if amount is None:
                os.remove(file_path)
                detail = "Сума не може бути пустою"
                raise HTTPException(status_code=400, detail=detail)

            if Plans.check_if_exists(month.strftime('%d.%m.%Y'), category_id):
                os.remove(file_path)
                detail = "План з таким місяцем та категорією вже існує"
                raise HTTPException(status_code=409, detail=detail)
            plan = Plans(period=month.strftime('%d.%m.%Y'), category_id=category_id, sum=amount)
            session.add(plan)

        session.commit()


        os.remove(file_path)

        return {"message": "Дані успішно внесено до БД"}
    except Exception as e:
        return {"error": str(e)}


@app.get("/plans_performance/")
async def get_plans_performance(date: str = Query()):
    try:
        date = datetime.strptime(date, "%d.%m.%Y")
        period = date.replace(day=1)
        plans = Plans.get_all_by_period(period)
        if not plans:
            print("Для цього місяця немає плану")
            raise HTTPException(status_code=404, detail="Для цього місяця немає плану")

        result = []
        for plan in plans:
            category = plan.category

            plan_info = {
                "Місяць плану": plan.period,
                "Категорія плану": category.name,
                "Сума з плану": plan.sum,
            }
            if category.name == "видача":
                amount = Credits.get_sum_by_date(period, date)
                plan_info["Сума виданих кредитів"] = amount
            elif category.name == "збір":
                amount = Payments.get_sum_by_date(period, date)
                plan_info["Сума платежів"] = amount
            else:
                amount = 0
            plan_info["% виконання плану"] = f"{(amount / plan.sum) * 100}" if plan.sum else "0%"
            result.append(plan_info)

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/year_performance/")
def get_summary_for_year(year: int = Query()):
    if not isinstance(year, int) or year <= 0:
        raise HTTPException(status_code=400, detail="Невірно введено рік, будь ласка введіть додатнє число.")
    credit_type_id = Dictionary.get_id_by_name("видача")
    payment_type_id = Dictionary.get_id_by_name("збір")
    payments_sum_for_year = Payments.get_sum_by_year(year)
    credit_sum_for_year = Credits.get_sum_by_year(year)
    summary = []
    for month in range(1, 13):
        credit_plan = Plans.get_by_category_and_month(month, year, credit_type_id)
        payment_plan = Plans.get_by_category_and_month(month, year, payment_type_id)
        credit_sum = Credits.get_sum_by_month(month, year)
        payment_sum = Payments.get_sum_by_month(month, year)
        month_summary = {
            "Місяць": month,
            "Рік": year,
            "Кількість видач за місяць": Credits.get_quantity_by_month(month, year),
            "Сума з плану по видачам на місяць": credit_plan.sum,
            "Сума видач за місяць": credit_sum,
            "% виконання плану по видачам": f"{credit_sum / credit_plan.sum * 100}%",
            "Кількість платежів за місяць": Payments.get_quantity_by_month(month, year),
            "Сума з плану по збору за місяць": payment_plan.sum,
            "Сума платежів за місяць": payment_sum,
            "% виконання плану по збору": f"{payment_sum / payment_plan.sum * 100}%",
            "% суми видач за місяць від суми видач за рік": f"{credit_sum / credit_sum_for_year * 100}%",
            "% суми платежів за місяць від суми платежів за рік": f"{payment_sum / payments_sum_for_year * 100}%"
        }
        summary.append(month_summary)

    return summary
